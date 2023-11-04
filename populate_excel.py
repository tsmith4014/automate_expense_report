
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import os

def populate_template(data, template_path, output_path):
    """
    Populates an Excel template with data and saves it to a new file.

    Args:
        data (dict): A dictionary containing the data to be populated in the template.
        template_path (str): The file path of the Excel template to be loaded.
        output_path (str): The file path of the new Excel file to be saved.

    Returns:
        None
    """

    # Load the Excel template
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active

    # Define the center alignment style with text wrapping
    center_aligned_text = Alignment(horizontal='center', wrapText=True)

    # Insert and resize the image
    img_path = os.path.join(os.path.dirname(__file__), 'ea.jpg')
    img = Image(img_path)
    # Set the size of the image
    img.width, img.height = img.width * 0.43, img.height * 0.60  # Resize to 50% of the original size
    sheet.add_image(img, 'A1')

    # Populate the template with data
    # Apply center alignment to the cells and set the values
    school_cell = sheet['B5']
    school_cell.alignment = center_aligned_text
    school_cell.value = data['school']

    period_ending_cell = sheet['H4']
    period_ending_cell.alignment = center_aligned_text
    period_ending_cell.value = datetime.strptime(data['period_ending'], '%Y-%m-%d').date()

    trip_purpose_cell = sheet['H5']
    trip_purpose_cell.alignment = center_aligned_text
    trip_purpose_cell.value = data['trip_purpose']

    # Calculate and populate dates based on the period ending date
    period_ending_date = datetime.strptime(data['period_ending'], '%Y-%m-%d')

    # Populate static day names with "Date" before each day
    day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for i, day_name in enumerate(day_names):
        day_cell = sheet.cell(row=7, column=4+i)
        day_cell.value = f"Date\n{day_name}"
        day_cell.alignment = center_aligned_text

    # Populate dates starting from the period ending date (Saturday) and going backward
    for i in range(6, -1, -1):  # Start from Saturday and go back to Sunday
        date_cell = sheet.cell(row=8, column=4+i)
        date_cell.value = (period_ending_date - timedelta(days=6-i)).date()
        date_cell.alignment = center_aligned_text

    # Save the populated template to a new file
    wb.save(output_path)
